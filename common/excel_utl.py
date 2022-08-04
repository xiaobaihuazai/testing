# -*- coding: utf-8 -*-
# @Time : 2022/7/31 15:45
# @Author : Aries
# @Site : 
# @File : excel_utl.py
# @Software: PyCharm
import os

import openpyxl as openpyxl


class ExcelUtl:

    def get_object_path(self):
        return os.path.dirname(__file__).split("common")[0]

    def read_excel(self):
        # 加载excel工作簿
        fil = openpyxl.load_workbook(self.get_object_path() + "date/dates.xlsx")
        # 获取sheet对象
        sheet = fil["Sheet1"]
        # 获取最大行与最大列
        print(sheet.max_row, sheet.max_column)
        all_list = []
        for rows in range(2, sheet.max_row+1):
            tmep_list = []
            for cols in range(1, sheet.max_column+1):
                # print(rows, cols)
                tmep_list.append(sheet.cell(rows, cols).value)
            all_list.append(tmep_list)
        return all_list


if __name__ == '__main__':
    print(ExcelUtl().read_excel())
