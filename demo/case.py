# -*- coding: utf-8 -*-
# @Time : 2022/8/1 20:05
# @Author : Aries
# @Site : 
# @File : case.py
# @Software: PyCharm

"""
create table 表名 (

    字段名 字段类型 comment '字段的注释'

) comment='表注释';
"""
import os
import openpyxl


class ExcelUtl:

    def get_object_path(self):
        return os.path.dirname(__file__).split("testcase")[0]

    def read_excel(self):
        # 加载excel工作簿
        fil = openpyxl.load_workbook(self.get_object_path() + "date/date2.xlsx")
        # 获取sheet对象
        sheet = fil["Sheet1"]
        # 获取最大行与最大列
        print(sheet.max_row, sheet.max_column)
        all_list = []
        for rows in range(2, sheet.max_row+1):
            tmep_list = []
            for cols in range(1, sheet.max_column+1):
                # print(rows, cols)
                tmep_list.append(sheet.cell(rows, cols).value)
            all_list.append(tmep_list)
        return all_list

    def csrt_create(self, 表名, 表注释):
        stes = self.read_excel()
        # print(stes)
        print("+++++++++++++++++++++++++++++++++++")

        zd1 = ""
        zd2 = ""
        sql = "create table 表名 ( " + zd2 + ") comment='表注释';"

        for i in range(0, len(stes)):

            if stes[i][1] is not None:
                s1 = stes[i][1]
                s2 = stes[i][2]
                if stes[i][3] == "字符":
                    s3 = "VARCHAR"
                elif stes[i][3] == "日期":
                    s3 = "DATE"
                elif stes[i][3] == "数字":
                    if isinstance(stes[i][4], int):
                        s3 = "INT"
                    else:
                        s3 = "DOUBLE"
                s4 = stes[i][4]
                if stes[i][5] == "是":
                    s5 = "not null"
                else:
                    s5 = ""
                # if stes[i][6] is not None:
                #     s6 = stes[i][6]
                # else:
                #     s6 = ""
                if s3 == "DATE":
                    zd = f" {s2} {s3} {s5} comment '{s1}'"
                else:
                    zd = f" {s2} {s3}({s4}) {s5} comment '{s1}'"

                if i == len(stes)-1:
                    zd1 = zd1 + zd + "\n"
                else:
                    zd1 = zd1 + zd + ",\n"
        # 表名 = "MDCS_FUND_SETL_LIST_D"
        # 表注释 = "医疗保障基金结算清单信息表（主表）"
        print("create table " + 表名 + " (" + zd1 + ") comment='" + 表注释 + "';")

    def csrt_select(self):
        stes = self.read_excel()
        ss = ""
        for i in range(0, len(stes)):
            if stes[i][1] is not None:
                if i == len(stes)-1:
                    st = stes[i][2] + "    #" + stes[i][1] + "\n"
                    # print(st)
                else:
                    st = stes[i][2] + ",    #" + stes[i][1] + "\n"
                    # print(st)
                ss = ss + st
        print("select \n" + ss + " FROM test_20220801;")

    def csrt_select2(self):
        stes = self.read_excel()
        ss = ""
        for i in range(0, len(stes)):
            if stes[i][1] is not None:
                if i == len(stes)-1:
                    st =stes[i][1] + " as " + stes[i][2] + "    #" + stes[i][1] + "\n"
                    # print(st)
                else:
                    st =stes[i][1] + " as " + stes[i][2] + ",    #" + stes[i][1] + "\n"
                    # print(st)
                ss = ss + st
        print("select \n" + ss + " FROM test_20220801;")

    def csrt_select3(self):
        stes = self.read_excel()
        ss = ""
        for i in range(0, len(stes)):
            if stes[i][1] is not None:
                if i == len(stes)-1:
                    st = stes[i][2] + " " + "\n"
                    # print(st)
                else:
                    st = stes[i][2] + ", " + "\n"
                    # print(st)
                ss = ss + st
        print("select \n" + ss + " FROM test_20220801;")

if __name__ == '__main__':
    # print(ExcelUtl().csrt_create("MDCS_FUND_SETL_LIST_DIAG_D", "2.2医疗保障基金结算清单诊断信息表"))
    # print(ExcelUtl().csrt_select())
    # print(ExcelUtl().csrt_select2())
    print(ExcelUtl().csrt_select3())